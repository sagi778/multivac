import asyncio
import datetime
import decimal
import json
import os
import warnings
from pathlib import Path
from typing import Any

import google.auth
from dotenv import load_dotenv
from fastapi import FastAPI, HTTPException, Request
from fastapi.responses import FileResponse
from fastapi.staticfiles import StaticFiles
from google.cloud import bigquery
from pydantic import BaseModel

load_dotenv()

DEFAULT_PROJECT = os.getenv("BQ_DEFAULT_PROJECT", "rapyd-eu-data")
MAX_ROWS = int(os.getenv("QUERY_MAX_ROWS", "1000"))

app = FastAPI(title="MultiVac")
STATIC_DIR = Path(__file__).parent / "static"
SESSION_FILE = Path(__file__).parent / "session.json"


def make_client(project: str = DEFAULT_PROJECT) -> bigquery.Client:
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        creds, _ = google.auth.default()
    return bigquery.Client(project=project, credentials=creds)


def stringify_value(v: Any) -> Any:
    if v is None:
        return None
    if isinstance(v, (datetime.datetime, datetime.date, datetime.time)):
        return v.isoformat()
    if isinstance(v, decimal.Decimal):
        return str(v)
    if isinstance(v, bytes):
        return v.hex()
    if isinstance(v, (dict, list)):
        return json.dumps(v, default=str)
    return v


def to_df_value(v: Any) -> Any:
    """Like stringify_value but converts Decimal to float so Polars infers numeric dtype.
    Used for story-block DataFrame creation; SUM/AVG on NUMERIC columns return Decimal."""
    if isinstance(v, decimal.Decimal):
        return float(v)
    return stringify_value(v)


def _make_df(col_data: dict) -> Any:
    """Build a Polars DataFrame from a {col: [values]} dict using strict=False so that
    columns with mixed int/float values (e.g. NUMERIC columns round-tripped through JSON)
    are cast to Float64 instead of raising a type-mismatch error."""
    import polars as pl
    if not col_data:
        return pl.DataFrame()
    series = [pl.Series(col, vals, strict=False) for col, vals in col_data.items()]
    return pl.DataFrame(series)


# ─── BQ helper functions (user-provided pattern) ─────────────────────────────

def get_bq_projects(project_name: str = DEFAULT_PROJECT) -> list[str]:
    client = make_client(project_name)
    projects = client.list_projects()
    return [p.project_id for p in projects]


def get_tables(project_id: str = DEFAULT_PROJECT, dataset_id: str | None = None) -> list[str]:
    client = make_client(project_id)
    tables_list = []
    if dataset_id is None:
        for dataset in client.list_datasets(project=project_id):
            ds_id = dataset.dataset_id
            for table in client.list_tables(f"{project_id}.{ds_id}"):
                tables_list.append(f"{project_id}.{ds_id}.{table.table_id}")
    else:
        for table in client.list_tables(f"{project_id}.{dataset_id}"):
            tables_list.append(f"{project_id}.{dataset_id}.{table.table_id}")
    return tables_list


# ─── Routes ──────────────────────────────────────────────────────────────────

@app.get("/", include_in_schema=False)
async def root():
    return FileResponse(STATIC_DIR / "index.html")


@app.post("/api/session/save")
async def save_session(request: Request):
    try:
        body = await request.body()
        SESSION_FILE.write_bytes(body)
        return {"ok": True}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/session/load")
async def load_session():
    if not SESSION_FILE.exists():
        return {"workspaces": None}
    try:
        return json.loads(SESSION_FILE.read_text(encoding="utf-8"))
    except Exception:
        return {"workspaces": None}


@app.get("/api/projects")
async def list_projects():
    try:
        project_ids = await asyncio.to_thread(get_bq_projects)
        return [{"id": pid, "displayName": pid} for pid in project_ids]
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@app.get("/api/projects/{project}/datasets")
async def list_datasets(project: str):
    try:
        client = make_client(project)
        datasets = await asyncio.to_thread(lambda: list(client.list_datasets(project=project)))
        return [{"id": ds.dataset_id, "project": ds.project} for ds in datasets]
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@app.get("/api/datasets/{project}/{dataset}/tables")
async def list_tables_endpoint(project: str, dataset: str):
    try:
        client = make_client(project)
        tables = await asyncio.to_thread(lambda: list(client.list_tables(f"{project}.{dataset}")))
        return [
            {"id": t.table_id, "project": t.project, "dataset": t.dataset_id, "tableType": t.table_type}
            for t in tables
        ]
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@app.get("/api/projects/{project}/tables/search")
async def search_tables(project: str, q: str = "", limit: int = 50):
    try:
        def _search():
            client = make_client(project)
            results = []
            q_lower = q.strip().lower()
            for dataset in client.list_datasets(project=project):
                ds_id = dataset.dataset_id
                ds_matches = not q_lower or q_lower in ds_id.lower()
                for table in client.list_tables(f"{project}.{ds_id}"):
                    if ds_matches or q_lower in table.table_id.lower():
                        results.append({
                            "id": table.table_id,
                            "project": project,
                            "dataset": ds_id,
                            "tableType": table.table_type,
                        })
                        if len(results) >= limit:
                            return results
            return results
        return await asyncio.to_thread(_search)
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@app.get("/api/tables/{project}/{dataset}/{table}/schema")
async def get_schema(project: str, dataset: str, table: str):
    try:
        client = make_client(project)
        ref = await asyncio.to_thread(client.get_table, f"{project}.{dataset}.{table}")
        return {
            "tableDescription": ref.description or "",
            "columns": [
                {"name": f.name, "type": f.field_type, "mode": f.mode, "description": f.description or ""}
                for f in ref.schema
            ],
        }
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


class QueryBody(BaseModel):
    sql: str
    project: str = DEFAULT_PROJECT
    limit_rows: int | None = None


@app.post("/api/query")
async def run_query(body: QueryBody):
    try:
        client = make_client(body.project or DEFAULT_PROJECT)
        job = client.query(body.sql)
        result = await asyncio.to_thread(job.result, max_results=body.limit_rows)
        columns = [field.name for field in result.schema]
        schema  = [{"name": f.name, "type": f.field_type} for f in result.schema]
        rows = [[stringify_value(v) for v in row.values()] for row in result]
        return {"columns": columns, "schema": schema, "rows": rows, "totalRows": result.total_rows, "jobId": job.job_id}
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


class LocalQueryBody(BaseModel):
    sql: str
    columns: list[str]
    rows: list[list]
    bq_schema: list[dict] = []
    extra_tables: dict = {}  # {"table_name": {"columns": [...], "rows": [...], "bq_schema": [...]}}


# BQ type → Polars dtype (used to re-cast string-serialised temporal/bool columns)
_BQ_TO_POLARS: dict = {}


def _get_bq_polars_map():
    global _BQ_TO_POLARS
    if _BQ_TO_POLARS:
        return _BQ_TO_POLARS
    import polars as pl
    _BQ_TO_POLARS = {
        "DATE":        ("date",     lambda s: s.str.to_date(strict=False)),
        "DATETIME":    ("datetime", lambda s: s.str.to_datetime(strict=False)),
        "TIMESTAMP":   ("datetime", lambda s: s.str.to_datetime(time_unit="us", time_zone="UTC", strict=False)),
        "TIME":        ("time",     lambda s: s.str.to_time(strict=False)),
        "INT64":       ("int",      lambda s: s.cast(pl.Int64, strict=False)),
        "INTEGER":     ("int",      lambda s: s.cast(pl.Int64, strict=False)),
        "INT":         ("int",      lambda s: s.cast(pl.Int64, strict=False)),
        "FLOAT64":     ("float",    lambda s: s.cast(pl.Float64, strict=False)),
        "FLOAT":       ("float",    lambda s: s.cast(pl.Float64, strict=False)),
        "NUMERIC":     ("float",    lambda s: s.cast(pl.Float64, strict=False)),
        "BIGNUMERIC":  ("float",    lambda s: s.cast(pl.Float64, strict=False)),
        "DECIMAL":     ("float",    lambda s: s.cast(pl.Float64, strict=False)),
        "BOOL":        ("bool",     lambda s: s.cast(pl.Boolean, strict=False)),
        "BOOLEAN":     ("bool",     lambda s: s.cast(pl.Boolean, strict=False)),
    }
    return _BQ_TO_POLARS


import re as _re
_ISO_DATE_RE     = _re.compile(r'^\d{4}-\d{2}-\d{2}$')
_ISO_DATETIME_RE = _re.compile(r'^\d{4}-\d{2}-\d{2}[T ]\d{2}:\d{2}')
_ISO_TZ_RE       = _re.compile(r'([+-]\d{2}:\d{2}|Z|\s+UTC)$')


def _auto_cast_temporal(df):
    """Cast String columns whose values all look like ISO dates/datetimes to Date/Datetime."""
    import polars as pl
    exprs = []
    for col in df.columns:
        if df[col].dtype != pl.String:
            continue
        sample = df[col].drop_nulls().head(50).to_list()
        if not sample:
            continue
        if all(_ISO_DATETIME_RE.match(str(v)) for v in sample):
            has_tz = any(_ISO_TZ_RE.search(str(v)) for v in sample[:5])
            if has_tz:
                exprs.append(pl.col(col).str.to_datetime(time_unit="us", time_zone="UTC", strict=False).alias(col))
            else:
                exprs.append(pl.col(col).str.to_datetime(strict=False).alias(col))
        elif all(_ISO_DATE_RE.match(str(v)) for v in sample):
            exprs.append(pl.col(col).str.to_date(strict=False).alias(col))
    return df.with_columns(exprs) if exprs else df


def _cast_df_by_bq_schema(df, bq_schema: list[dict]):
    """Cast string columns in a Polars DataFrame back to their original BQ types.
    Also auto-detects STRING columns whose values are ISO date/datetime strings
    (e.g. FORMAT_DATE output) and promotes them to Date/Datetime."""
    import polars as pl
    type_map = {s["name"]: s["type"].upper().split("<")[0].split("[")[0] for s in (bq_schema or [])}
    casters = _get_bq_polars_map()
    exprs = []
    for col in df.columns:
        if df[col].dtype != pl.String:
            continue
        bq_type = type_map.get(col, "")
        if bq_type in casters:
            _, cast_fn = casters[bq_type]
            try:
                exprs.append(cast_fn(pl.col(col)).alias(col))
            except Exception:
                pass
    if exprs:
        try:
            df = df.with_columns(exprs)
        except Exception:
            # Apply casts one-by-one so a bad column doesn't block the rest
            for expr in exprs:
                try:
                    df = df.with_columns([expr])
                except Exception:
                    pass
    # Auto-promote remaining String columns that look like ISO dates
    return _auto_cast_temporal(df)


@app.post("/api/query/local")
async def run_local_query(body: LocalQueryBody):
    def _run():
        import polars as pl
        col_data = {col: [to_df_value(body.rows[r][i]) for r in range(len(body.rows))]
                    for i, col in enumerate(body.columns)}
        df = _make_df(col_data) if body.rows else pl.DataFrame({c: [] for c in body.columns})
        df = _cast_df_by_bq_schema(df, body.bq_schema)
        tables = {"df": df}
        for tname, tdata in (body.extra_tables or {}).items():
            tc = {col: [to_df_value(tdata['rows'][r][i]) for r in range(len(tdata['rows']))]
                  for i, col in enumerate(tdata['columns'])}
            t_df = _make_df(tc) if tdata['rows'] else pl.DataFrame({c: [] for c in tdata['columns']})
            t_df = _cast_df_by_bq_schema(t_df, tdata.get('bq_schema', []))
            tables[tname] = t_df
        ctx = pl.SQLContext(tables)
        out = ctx.execute(body.sql).collect()
        schema = [{"name": c, "type": str(out.schema[c])} for c in out.columns]
        return {
            "columns": out.columns,
            "schema": schema,
            "rows": [[stringify_value(v) for v in row] for row in out.rows()],
            "totalRows": out.height,
        }
    try:
        return await asyncio.to_thread(_run)
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


class StoryBlockRequest(BaseModel):
    sql: str
    project: str = DEFAULT_PROJECT
    block_type: str
    params: dict = {}
    # Optional pre-loaded data (skips BQ query when provided)
    local_columns: list = []
    local_rows: list = []
    local_bq_schema: list = []
    local_extra_tables: dict = {}  # extra sheets for Excel workspaces


@app.post("/api/story/block")
async def run_story_block(body: StoryBlockRequest):
    def _run():
        import polars as pl
        if body.local_columns:
            # Use pre-loaded data (data file workspace — skip BQ query)
            col_data = {col: [to_df_value(body.local_rows[r][i]) for r in range(len(body.local_rows))]
                        for i, col in enumerate(body.local_columns)}
            df = _make_df(col_data) if body.local_rows else pl.DataFrame({c: [] for c in body.local_columns})
            df = _cast_df_by_bq_schema(df, body.local_bq_schema)
            columns = body.local_columns
        else:
            client = make_client(body.project or DEFAULT_PROJECT)
            result = client.query(body.sql).result()
            columns = [f.name for f in result.schema]
            raw = [[to_df_value(v) for v in row.values()] for row in result]
            df = _make_df({col: [r[i] for r in raw] for i, col in enumerate(columns)}) if raw \
                 else pl.DataFrame({col: [] for col in columns})

        # Register extra tables (Excel sheets) in module-level dict for sql_query blocks if needed
        _extra_frames = {}
        for tname, tdata in (body.local_extra_tables or {}).items():
            tc = {col: [to_df_value(tdata['rows'][r][i]) for r in range(len(tdata['rows']))]
                  for i, col in enumerate(tdata['columns'])}
            t_df = _make_df(tc) if tdata['rows'] else pl.DataFrame({c: [] for c in tdata['columns']})
            _extra_frames[tname] = _cast_df_by_bq_schema(t_df, tdata.get('bq_schema', []))

        if body.block_type == "shape":
            return {"rows": df.height, "columns": len(df.columns)}

        if body.block_type == "preview":
            cols   = [c for c in body.params.get("columns", []) if c in df.columns]
            if cols:
                df = df.select(cols)
            amount = max(1, int(body.params.get("amount", 5)))
            mode   = body.params.get("mode", "head")
            sub = df.tail(amount) if mode == "tail" \
                  else df.sample(min(amount, df.height), shuffle=True) if mode == "random" \
                  else df.head(amount)
            _int_dt   = {pl.Int8, pl.Int16, pl.Int32, pl.Int64, pl.UInt8, pl.UInt16, pl.UInt32, pl.UInt64}
            _float_dt = {pl.Float32, pl.Float64}
            def _bq_type(dt):
                if dt in _int_dt:   return "INTEGER"
                if dt in _float_dt: return "NUMERIC"
                if dt == pl.Boolean: return "BOOL"
                if dt == pl.Date:   return "DATE"
                if isinstance(dt, pl.Datetime): return "TIMESTAMP"
                return "STRING"
            schema = [{"name": c, "type": _bq_type(sub[c].dtype)} for c in sub.columns]
            return {"columns": sub.columns, "rows": [list(r) for r in sub.iter_rows()], "schema": schema}

        if body.block_type == "search":
            _num_dt = {pl.Int8, pl.Int16, pl.Int32, pl.Int64,
                       pl.UInt8, pl.UInt16, pl.UInt32, pl.UInt64,
                       pl.Float32, pl.Float64}
            _date_dt = {pl.Date, pl.Datetime, pl.Duration, pl.Time}

            col_types = {}
            for c in df.columns:
                dt = df[c].dtype
                if dt in _num_dt:
                    col_types[c] = "numeric"
                elif dt in _date_dt:
                    col_types[c] = "date"
                else:
                    col_types[c] = "string"

            filters = body.params.get("filters", {})
            for col_name, f in filters.items():
                if col_name not in df.columns or not isinstance(f, dict):
                    continue
                try:
                    ctype = col_types.get(col_name, "string")
                    if ctype == "numeric":
                        mn, mx = f.get("min"), f.get("max")
                        if mn is not None and mn != "":
                            df = df.filter(pl.col(col_name) >= float(mn))
                        if mx is not None and mx != "":
                            df = df.filter(pl.col(col_name) <= float(mx))
                    elif ctype == "date":
                        frm, to = f.get("from"), f.get("to")
                        if frm:
                            df = df.filter(pl.col(col_name).cast(pl.Utf8) >= str(frm))
                        if to:
                            df = df.filter(pl.col(col_name).cast(pl.Utf8) <= str(to))
                    else:
                        contains = (f.get("contains") or "").strip()
                        op = f.get("op", "contains")
                        if contains:
                            col_str = pl.col(col_name).cast(pl.Utf8)
                            if op == "equals":
                                df = df.filter(col_str == contains)
                            elif op == "starts_with":
                                df = df.filter(col_str.str.starts_with(contains))
                            elif op == "not_contains":
                                df = df.filter(~col_str.str.contains(contains, literal=True))
                            else:
                                df = df.filter(col_str.str.contains(contains, literal=True))
                except Exception:
                    pass

            limit = max(1, min(5000, int(body.params.get("rows", 50))))
            total = df.height
            sub   = df.head(limit)
            return {"columns": sub.columns, "col_types": col_types,
                    "total": total, "showing": sub.height,
                    "rows": [list(r) for r in sub.iter_rows()]}

        if body.block_type == "describe":
            total = df.height
            rows = []
            for col_name in df.columns:
                col = df[col_name]
                null_count = col.null_count()
                null_pct = round(null_count / total * 100, 2) if total > 0 else 0.0
                non_null = col.drop_nulls()
                samples = []
                if non_null.len() > 0:
                    try:
                        vc = non_null.value_counts().sort("count", descending=True).head(5)
                        samples = [{"value": str(r[0]), "count": int(r[1])} for r in vc.iter_rows()]
                    except Exception:
                        pass
                rows.append({
                    "column": col_name,
                    "dtype": str(col.dtype),
                    "nulls": null_count,
                    "null_pct": null_pct,
                    "samples": samples,
                })
            return {"describe": rows, "total_rows": total}

        if body.block_type == "histogram":
            col_name = body.params.get("column", "")
            if not col_name or col_name not in df.columns:
                col_name = df.columns[0] if df.columns else None
            if not col_name:
                raise HTTPException(status_code=400, detail="No column to plot")
            col = df[col_name].drop_nulls()
            bins = max(2, min(100, int(body.params.get("bins", 20))))
            numeric_dtypes = {
                pl.Int8, pl.Int16, pl.Int32, pl.Int64,
                pl.UInt8, pl.UInt16, pl.UInt32, pl.UInt64,
                pl.Float32, pl.Float64,
            }
            if col.dtype in numeric_dtypes:
                if col.len() == 0:
                    return {"histogram": [], "column": col_name, "kind": "numeric"}
                min_v = float(col.min())
                max_v = float(col.max())
                if min_v == max_v:
                    return {"histogram": [{"label": str(min_v), "count": int(col.len())}], "column": col_name, "kind": "numeric"}
                bw = (max_v - min_v) / bins
                result = []
                for i in range(bins):
                    lo = min_v + i * bw
                    hi = min_v + (i + 1) * bw
                    mask = (col >= lo) & (col < hi) if i < bins - 1 else (col >= lo) & (col <= hi)
                    result.append({"label": f"{lo:.3g}", "count": int(col.filter(mask).len())})
                return {"histogram": result, "column": col_name, "kind": "numeric"}
            else:
                vc = col.value_counts().sort("count", descending=True).head(30)
                result = [{"label": str(r[0]), "count": int(r[1])} for r in vc.iter_rows()]
                return {"histogram": result, "column": col_name, "kind": "categorical"}

        if body.block_type == "col_describe":
            col_name = body.params.get("column", "")
            if not col_name or col_name not in df.columns:
                raise HTTPException(status_code=400, detail="Select a column")
            col = df[col_name]
            total = df.height
            null_count = col.null_count()
            non_null = total - null_count
            null_pct = round(null_count / total * 100, 2) if total > 0 else 0.0
            col_clean = col.drop_nulls()
            result = {
                "column": col_name,
                "dtype": str(col.dtype),
                "total": total,
                "non_null": non_null,
                "nulls": null_count,
                "null_pct": null_pct,
            }
            numeric_dtypes = {
                pl.Int8, pl.Int16, pl.Int32, pl.Int64,
                pl.UInt8, pl.UInt16, pl.UInt32, pl.UInt64,
                pl.Float32, pl.Float64,
            }
            if col.dtype in numeric_dtypes and col_clean.len() > 0:
                import numpy as np
                col_min  = float(col_clean.min())
                col_max  = float(col_clean.max())
                col_mean = round(float(col_clean.mean()), 6)
                col_std  = round(float(col_clean.std()), 6)
                q25      = float(col_clean.quantile(0.25))
                q50      = float(col_clean.quantile(0.5))
                q75      = float(col_clean.quantile(0.75))
                col_range = col_max - col_min
                bias = round((col_mean - q50) / col_range, 6) if col_range != 0 else 0.0
                # histogram for distribution chart
                n_bins = min(40, max(10, int(col_clean.len() ** 0.45)))
                values_np = col_clean.to_numpy()
                hist_counts, hist_edges = np.histogram(values_np, bins=n_bins)
                iqr = q75 - q25
                lo_fence = q25 - 1.5 * iqr
                hi_fence = q75 + 1.5 * iqr
                whisker_lo = float(col_clean.filter(col_clean >= lo_fence).min() or lo_fence)
                whisker_hi = float(col_clean.filter(col_clean <= hi_fence).max() or hi_fence)
                result.update({
                    "kind": "numeric",
                    "min": col_min,
                    "max": col_max,
                    "mean": col_mean,
                    "std": col_std,
                    "q25": q25,
                    "q50": q50,
                    "q75": q75,
                    "bias": bias,
                    "whisker_lo": whisker_lo,
                    "whisker_hi": whisker_hi,
                    "histogram": [
                        {"x0": float(hist_edges[i]), "x1": float(hist_edges[i+1]),
                         "count": int(hist_counts[i])}
                        for i in range(len(hist_counts))
                    ],
                })
            else:
                vc = col_clean.value_counts().sort("count", descending=True).head(10)
                result.update({
                    "kind": "categorical",
                    "distinct": int(col_clean.n_unique()),
                    "top": [{"value": str(r[0]), "count": int(r[1])} for r in vc.iter_rows()],
                })
            return result

        if body.block_type == "scatter":
            x_col = body.params.get("x_col", "")
            y_col = body.params.get("y_col", "")
            group_by = body.params.get("group_by", "")
            if not x_col or x_col not in df.columns:
                raise HTTPException(status_code=400, detail="Select X column")
            if not y_col or y_col not in df.columns:
                raise HTTPException(status_code=400, detail="Select Y column")
            max_pts = max(10, min(2000, int(body.params.get("max_points", 500))))
            if group_by and group_by in df.columns:
                groups = df[group_by].drop_nulls().unique().sort().head(10).to_list()
                pts_per = max(10, max_pts // max(len(groups), 1))
                series = []
                for g in groups:
                    sub = df.filter(pl.col(group_by) == g).select([x_col, y_col]).drop_nulls()
                    if sub.height > pts_per:
                        sub = sub.sample(pts_per, shuffle=True)
                    series.append({
                        "label": str(g),
                        "x": [stringify_value(v) for v in sub[x_col].to_list()],
                        "y": [stringify_value(v) for v in sub[y_col].to_list()],
                    })
                return {"series": series, "x_col": x_col, "y_col": y_col, "grouped": True}
            else:
                sub = df.select([x_col, y_col]).drop_nulls()
                if sub.height > max_pts:
                    sub = sub.sample(max_pts, shuffle=True)
                return {
                    "x": [stringify_value(v) for v in sub[x_col].to_list()],
                    "y": [stringify_value(v) for v in sub[y_col].to_list()],
                    "x_col": x_col, "y_col": y_col, "grouped": False,
                }

        if body.block_type == "lineplot":
            x_col = body.params.get("x_col", "")
            y_col = body.params.get("y_col", "")
            group_by = body.params.get("group_by", "")
            if not x_col or x_col not in df.columns:
                raise HTTPException(status_code=400, detail="Select X column")
            if not y_col or y_col not in df.columns:
                raise HTTPException(status_code=400, detail="Select Y column")
            if group_by and group_by in df.columns:
                groups = df[group_by].drop_nulls().unique().sort().head(10).to_list()
                series = []
                for g in groups:
                    sub = df.filter(pl.col(group_by) == g).select([x_col, y_col]).drop_nulls()
                    try:
                        sub = sub.sort(x_col)
                    except Exception:
                        pass
                    if sub.height > 1000:
                        sub = sub.head(1000)
                    series.append({
                        "label": str(g),
                        "x": [stringify_value(v) for v in sub[x_col].to_list()],
                        "y": [stringify_value(v) for v in sub[y_col].to_list()],
                    })
                return {"series": series, "x_col": x_col, "y_col": y_col, "grouped": True}
            else:
                sub = df.select([x_col, y_col]).drop_nulls()
                try:
                    sub = sub.sort(x_col)
                except Exception:
                    pass
                if sub.height > 1000:
                    sub = sub.head(1000)
                return {
                    "x": [stringify_value(v) for v in sub[x_col].to_list()],
                    "y": [stringify_value(v) for v in sub[y_col].to_list()],
                    "x_col": x_col, "y_col": y_col, "grouped": False,
                }

        if body.block_type == "boxplot":
            y_col = body.params.get("y_col", "")
            group_col = body.params.get("group_col", "")
            if not y_col or y_col not in df.columns:
                raise HTTPException(status_code=400, detail="Select column")
            numeric_dtypes = {
                pl.Int8, pl.Int16, pl.Int32, pl.Int64,
                pl.UInt8, pl.UInt16, pl.UInt32, pl.UInt64,
                pl.Float32, pl.Float64,
            }
            if df[y_col].dtype not in numeric_dtypes:
                raise HTTPException(status_code=400, detail=f"Column '{y_col}' must be numeric for box plot")

            def compute_box(series):
                s = series.drop_nulls()
                if s.len() == 0:
                    return None
                q25 = float(s.quantile(0.25))
                q50 = float(s.quantile(0.5))
                q75 = float(s.quantile(0.75))
                iqr = q75 - q25
                lo_fence = q25 - 1.5 * iqr
                hi_fence = q75 + 1.5 * iqr
                whisker_lo = float(s.filter(s >= lo_fence).min())
                whisker_hi = float(s.filter(s <= hi_fence).max())
                mean_val = float(s.mean())
                # Return all points (capped at 150) for jittered dot plot
                points = [float(v) for v in (s.sample(150, shuffle=True) if s.len() > 150 else s).to_list()]
                return {"q25": q25, "q50": q50, "q75": q75, "lo": whisker_lo, "hi": whisker_hi,
                        "mean": mean_val, "points": points}

            if group_col and group_col in df.columns:
                groups = df[group_col].drop_nulls().unique().sort().head(12).to_list()
                boxes = []
                for g in groups:
                    b = compute_box(df.filter(pl.col(group_col) == g)[y_col])
                    if b:
                        b["label"] = str(g)
                        boxes.append(b)
            else:
                b = compute_box(df[y_col])
                boxes = [dict(b, label=y_col)] if b else []

            overall_series = df[y_col].drop_nulls()
            overall_mean = float(overall_series.mean()) if overall_series.len() > 0 else None
            return {"boxes": boxes, "y_col": y_col, "overall_mean": overall_mean}

        if body.block_type == "outlier_detect":
            col_name = body.params.get("column", "")
            method   = body.params.get("method", "IQR")
            threshold = float(body.params.get("threshold", 1.5))

            if not col_name or col_name not in df.columns:
                raise HTTPException(status_code=400, detail="Select a numeric column")
            try:
                col = df[col_name].cast(pl.Float64, strict=False)
            except Exception:
                raise HTTPException(status_code=400, detail=f"Column '{col_name}' must be numeric")

            n_null  = int(col.null_count())
            col     = col.drop_nulls()
            n_valid = col.len()
            if n_valid < 4:
                raise HTTPException(status_code=400, detail="Need at least 4 non-null values")

            values = col.to_list()
            lower = upper = 0.0
            stats: dict = {}

            if method == "IQR":
                q1 = float(col.quantile(0.25))
                q3 = float(col.quantile(0.75))
                iqr_v = q3 - q1
                lower = q1 - threshold * iqr_v
                upper = q3 + threshold * iqr_v
                flags = [v < lower or v > upper for v in values]
                stats = {"Q1": round(q1, 4), "Q3": round(q3, 4), "IQR": round(iqr_v, 4)}

            elif method == "Z-Score":
                mean = float(col.mean())
                std  = float(col.std())
                if std == 0:
                    flags = [False] * len(values)
                    lower = upper = mean
                else:
                    lower = mean - threshold * std
                    upper = mean + threshold * std
                    flags = [abs((v - mean) / std) > threshold for v in values]
                stats = {"Mean": round(mean, 4), "Std": round(std, 4)}

            elif method == "Mod-Z":
                median_v = float(col.median())
                mad = float((col - median_v).abs().median())
                if mad == 0:
                    flags = [False] * len(values)
                    lower = upper = median_v
                else:
                    lower = median_v - (threshold * mad / 0.6745)
                    upper = median_v + (threshold * mad / 0.6745)
                    flags = [(0.6745 * abs(v - median_v) / mad) > threshold for v in values]
                stats = {"Median": round(median_v, 4), "MAD": round(mad, 4)}

            elif method == "Percentile":
                pct = max(0.1, min(49.9, threshold))
                lower = float(col.quantile(pct / 100))
                upper = float(col.quantile(1 - pct / 100))
                flags = [v < lower or v > upper for v in values]
                stats = {"Low%": pct, "High%": round(100 - pct, 1)}

            else:
                raise HTTPException(status_code=400, detail=f"Unknown method: {method}")

            n_outliers = int(sum(flags))
            outlier_pct = round(n_outliers / n_valid * 100, 2) if n_valid > 0 else 0.0

            # Track original row indices of non-null values for column insertion
            non_null_idx = (
                df.with_row_index("__idx__")
                .filter(pl.col(col_name).cast(pl.Float64, strict=False).is_not_null())
                ["__idx__"].to_list()
            )

            # Sample chart points (always keep all outliers + random normal)
            import random as _rnd
            MAX_CHART = 600
            outlier_idx = [i for i, f in enumerate(flags) if f]
            normal_idx  = [i for i, f in enumerate(flags) if not f]
            keep_normal = _rnd.sample(normal_idx, min(len(normal_idx), MAX_CHART - len(outlier_idx)))
            keep = sorted(outlier_idx + keep_normal)
            chart_values = [round(float(values[i]), 6) for i in keep]
            chart_flags  = [flags[i] for i in keep]

            return {
                "column":        col_name,
                "method":        method,
                "threshold":     threshold,
                "n_outliers":    n_outliers,
                "n_valid":       n_valid,
                "n_null":        n_null,
                "outlier_pct":   outlier_pct,
                "lower":         round(lower, 6),
                "upper":         round(upper, 6),
                "chart_values":  chart_values,
                "chart_flags":   chart_flags,
                "stats":         stats,
                "sampled":       len(values) > MAX_CHART,
                "row_indices":   non_null_idx,
                "outlier_flags": [bool(f) for f in flags],
            }

        if body.block_type == "linreg":
            from scipy import stats as sp_stats
            import numpy as np
            x_col = body.params.get("x_col", "")
            y_col = body.params.get("y_col", "")
            group_col = (body.params.get("group_col") or "").strip()
            if not x_col or x_col not in df.columns:
                raise HTTPException(status_code=400, detail="Select X column")
            if not y_col or y_col not in df.columns:
                raise HTTPException(status_code=400, detail="Select Y column")
            try:
                x_s = df[x_col].cast(pl.Float64, strict=False)
                y_s = df[y_col].cast(pl.Float64, strict=False)
            except Exception:
                raise HTTPException(status_code=400, detail="Both columns must be numeric")
            max_pts = max(10, min(2000, int(body.params.get("max_points", 500))))

            # ── Grouped mode ──────────────────────────────────────────────────
            if group_col and group_col in df.columns:
                base = pl.DataFrame({"x": x_s, "y": y_s, "grp": df[group_col].cast(pl.String)}).drop_nulls(subset=["x", "y"])
                grp_vals = base["grp"].drop_nulls().unique().sort().to_list()
                if len(grp_vals) > 20:
                    grp_vals = grp_vals[:20]
                group_results = []
                per_grp_pts = max(10, max_pts // max(len(grp_vals), 1))
                for g in grp_vals:
                    g_sub = base.filter(pl.col("grp") == g)
                    if g_sub.height < 3:
                        continue
                    xa = g_sub["x"].to_numpy()
                    ya = g_sub["y"].to_numpy()
                    ng = len(xa)
                    try:
                        lr_g = sp_stats.linregress(xa, ya)
                    except Exception:
                        continue
                    sl = float(lr_g.slope)
                    if np.isnan(sl) or np.isinf(sl):
                        continue
                    ic = float(lr_g.intercept)
                    rv = float(lr_g.rvalue)
                    pv = float(lr_g.pvalue)
                    se = float(lr_g.stderr)
                    xmn, xmx = float(xa.min()), float(xa.max())
                    lx = [xmn + i * (xmx - xmn) / 49 for i in range(50)]
                    ly = [sl * xi + ic for xi in lx]
                    if ng > per_grp_pts:
                        idx = np.random.choice(ng, per_grp_pts, replace=False)
                        sc_x_g = xa[idx].tolist(); sc_y_g = ya[idx].tolist()
                    else:
                        sc_x_g = xa.tolist(); sc_y_g = ya.tolist()
                    group_results.append({
                        "group": str(g), "n": ng,
                        "slope": round(sl, 6), "intercept": round(ic, 6),
                        "r2": round(rv ** 2, 6), "r": round(rv, 6),
                        "p_value": pv, "std_err": round(se, 6),
                        "sc_x": [round(float(v), 6) for v in sc_x_g],
                        "sc_y": [round(float(v), 6) for v in sc_y_g],
                        "line_x": [round(v, 6) for v in lx],
                        "line_y": [round(v, 6) for v in ly],
                    })
                if not group_results:
                    raise HTTPException(status_code=400, detail="No groups had enough points (≥3) for regression")
                return {
                    "is_grouped": True,
                    "x_col": x_col, "y_col": y_col, "group_col": group_col,
                    "groups": group_results,
                }

            # ── Ungrouped mode ────────────────────────────────────────────────
            sub = pl.DataFrame({"x": x_s, "y": y_s}).drop_nulls()
            if sub.height < 3:
                raise HTTPException(status_code=400, detail="Need at least 3 non-null paired points")
            x_arr = sub["x"].to_numpy()
            y_arr = sub["y"].to_numpy()
            n = len(x_arr)
            try:
                lr = sp_stats.linregress(x_arr, y_arr)
            except Exception as e:
                raise HTTPException(status_code=400, detail=f"Cannot fit linear regression: {e}")
            slope = float(lr.slope)
            if np.isnan(slope) or np.isinf(slope):
                raise HTTPException(status_code=400, detail="Cannot fit linear regression (constant X or degenerate data)")
            intercept = float(lr.intercept)
            r_val = float(lr.rvalue)
            p_val = float(lr.pvalue)
            std_err = float(lr.stderr)
            r2 = r_val ** 2
            x_min, x_max = float(x_arr.min()), float(x_arr.max())
            line_x = [x_min + i * (x_max - x_min) / 99 for i in range(100)]
            line_y = [slope * xi + intercept for xi in line_x]
            if n > max_pts:
                idx = np.random.choice(n, max_pts, replace=False)
                sc_x = x_arr[idx].tolist()
                sc_y = y_arr[idx].tolist()
            else:
                sc_x = x_arr.tolist()
                sc_y = y_arr.tolist()
            show_ci = bool(body.params.get("show_ci", True))
            ci_lo = ci_hi = None
            if show_ci and n >= 3:
                x_mean = float(x_arr.mean())
                ss_x = float(((x_arr - x_mean) ** 2).sum())
                if ss_x > 0:
                    from scipy.stats import t as t_dist
                    se_fit = std_err * np.sqrt(1 / n + (np.array(line_x) - x_mean) ** 2 / ss_x)
                    t_crit = float(t_dist.ppf(0.975, df=n - 2))
                    ci_lo = (np.array(line_y) - t_crit * se_fit).tolist()
                    ci_hi = (np.array(line_y) + t_crit * se_fit).tolist()
            x_full = df[x_col].cast(pl.Float64, strict=False)
            pred_values = [
                round(slope * v + intercept, 6) if v is not None else None
                for v in x_full.to_list()
            ]
            return {
                "is_grouped": False,
                "x_col": x_col, "y_col": y_col, "n": n,
                "slope": round(slope, 6), "intercept": round(intercept, 6),
                "r2": round(r2, 6), "r": round(r_val, 6),
                "p_value": p_val, "std_err": round(std_err, 6),
                "sc_x": [round(float(v), 6) for v in sc_x],
                "sc_y": [round(float(v), 6) for v in sc_y],
                "line_x": [round(v, 6) for v in line_x],
                "line_y": [round(v, 6) for v in line_y],
                "ci_lo": [round(float(v), 6) for v in ci_lo] if ci_lo else None,
                "ci_hi": [round(float(v), 6) for v in ci_hi] if ci_hi else None,
                "sampled": n > max_pts,
                "pred_values": pred_values,
            }

        if body.block_type == "hypo_test":
            from scipy import stats as sp_stats
            import numpy as np
            value_col = body.params.get("value_col", "")
            group_col = body.params.get("group_col", "")
            test_mode = body.params.get("test", "auto")
            alpha = float(body.params.get("alpha", 0.05))
            if not value_col or value_col not in df.columns:
                raise HTTPException(status_code=400, detail="Select value column")
            if not group_col or group_col not in df.columns:
                raise HTTPException(status_code=400, detail="Select group column")
            try:
                val_s = df[value_col].cast(pl.Float64, strict=False)
            except Exception:
                raise HTTPException(status_code=400, detail="Value column must be numeric")
            sub = pl.DataFrame({"val": val_s, "grp": df[group_col]}).drop_nulls()
            groups = sub["grp"].unique().sort().to_list()
            n_groups = len(groups)
            if n_groups < 2:
                raise HTTPException(status_code=400, detail="Need at least 2 groups")
            if n_groups > 12:
                groups = groups[:12]
                sub = sub.filter(pl.col("grp").is_in(groups))
            group_arrays = [sub.filter(pl.col("grp") == g)["val"].to_numpy() for g in groups]
            # Resolve test type
            effective_test = test_mode
            if test_mode == "auto":
                effective_test = "anova" if n_groups > 2 else "t-test"
            elif test_mode in ("t-test", "mann-whitney") and n_groups > 2:
                effective_test = "anova"
            if effective_test == "t-test":
                r = sp_stats.ttest_ind(group_arrays[0], group_arrays[1], equal_var=False)
            elif effective_test == "mann-whitney":
                r = sp_stats.mannwhitneyu(group_arrays[0], group_arrays[1], alternative="two-sided")
            elif effective_test == "kruskal":
                r = sp_stats.kruskal(*group_arrays)
            else:
                r = sp_stats.f_oneway(*group_arrays)
                effective_test = "anova"
            statistic = float(r.statistic)
            p_value = float(r.pvalue)
            significant = bool(p_value < alpha)
            # Effect size
            effect_size = None
            effect_label = ""
            if effective_test in ("t-test", "mann-whitney") and n_groups == 2:
                n1, n2 = len(group_arrays[0]), len(group_arrays[1])
                if n1 > 1 and n2 > 1:
                    pooled_std = np.sqrt(
                        ((n1 - 1) * group_arrays[0].std() ** 2 + (n2 - 1) * group_arrays[1].std() ** 2)
                        / (n1 + n2 - 2)
                    )
                    if pooled_std > 0:
                        d = abs(group_arrays[0].mean() - group_arrays[1].mean()) / pooled_std
                        effect_size = round(float(d), 4)
                        effect_label = "small" if d < 0.5 else "medium" if d < 0.8 else "large"
            else:
                all_vals = np.concatenate(group_arrays)
                grand_mean = all_vals.mean()
                ss_between = sum(len(g) * (g.mean() - grand_mean) ** 2 for g in group_arrays)
                ss_total = ((all_vals - grand_mean) ** 2).sum()
                if ss_total > 0:
                    eta2 = ss_between / ss_total
                    effect_size = round(float(eta2), 4)
                    effect_label = "small" if eta2 < 0.06 else "medium" if eta2 < 0.14 else "large"
            # Group stats
            group_stats = []
            for g, arr in zip(groups, group_arrays):
                group_stats.append({
                    "group": str(g),
                    "n": int(len(arr)),
                    "mean": round(float(arr.mean()), 4) if len(arr) > 0 else None,
                    "std": round(float(arr.std(ddof=1)), 4) if len(arr) > 1 else None,
                    "median": round(float(np.median(arr)), 4) if len(arr) > 0 else None,
                })
            # Boxplot boxes
            def _box(arr):
                s = pl.Series(arr).drop_nulls()
                if s.len() == 0:
                    return None
                q25 = float(s.quantile(0.25))
                q50 = float(s.quantile(0.5))
                q75 = float(s.quantile(0.75))
                iqr_v = q75 - q25
                lo_fence = q25 - 1.5 * iqr_v
                hi_fence = q75 + 1.5 * iqr_v
                wlo = float(s.filter(s >= lo_fence).min())
                whi = float(s.filter(s <= hi_fence).max())
                pts = [float(v) for v in (s.sample(150, shuffle=True) if s.len() > 150 else s).to_list()]
                return {"q25": q25, "q50": q50, "q75": q75, "lo": wlo, "hi": whi,
                        "mean": float(s.mean()), "points": pts}
            boxes = []
            for g, arr in zip(groups, group_arrays):
                b = _box(arr)
                if b:
                    b["label"] = str(g)
                    boxes.append(b)
            # KDE distributions
            dists = []
            for g, arr in zip(groups, group_arrays):
                if len(arr) < 3:
                    continue
                kde = sp_stats.gaussian_kde(arr, bw_method="scott")
                xlo, xhi = float(arr.min()), float(arr.max())
                pad = (xhi - xlo) * 0.15 or 0.5
                xs = np.linspace(xlo - pad, xhi + pad, 100)
                ys = kde(xs)
                dists.append({"label": str(g), "x": xs.tolist(), "y": ys.tolist()})
            return {
                "value_col": value_col, "group_col": group_col,
                "test_used": effective_test,
                "statistic": round(statistic, 4),
                "p_value": p_value, "alpha": alpha,
                "significant": significant,
                "effect_size": effect_size, "effect_label": effect_label,
                "group_stats": group_stats,
                "boxes": boxes, "dists": dists,
            }

        if body.block_type == "barplot":
            x_col     = body.params.get("x_col", "")
            y_col     = body.params.get("y_col", "")
            agg       = body.params.get("agg", "count")
            group_col = body.params.get("group_col", "")
            sort      = body.params.get("sort", "desc")

            if not x_col or x_col not in df.columns:
                raise HTTPException(status_code=400, detail="Select a category (X) column")

            if agg != "count":
                if not y_col or y_col not in df.columns:
                    raise HTTPException(status_code=400, detail="Select a value (Y) column for the chosen aggregation")
                numeric_dtypes = (pl.Float64, pl.Float32, pl.Int64, pl.Int32, pl.Int16, pl.Int8,
                                  pl.UInt64, pl.UInt32, pl.UInt16, pl.UInt8)
                if df[y_col].dtype not in numeric_dtypes:
                    raise HTTPException(status_code=400, detail=f"Column '{y_col}' must be numeric for '{agg}' aggregation")
                agg_map = {
                    "sum":    pl.col(y_col).sum(),
                    "mean":   pl.col(y_col).mean(),
                    "median": pl.col(y_col).median(),
                    "min":    pl.col(y_col).min(),
                    "max":    pl.col(y_col).max(),
                }
                if agg not in agg_map:
                    raise HTTPException(status_code=400, detail=f"Unknown aggregation: {agg}")
                agg_expr = agg_map[agg].alias("value")
            else:
                agg_expr = pl.len().alias("value")

            use_group = bool(group_col and group_col in df.columns and group_col != x_col)
            group_keys = [x_col, group_col] if use_group else [x_col]
            result = df.group_by(group_keys).agg(agg_expr)

            y_label = agg if agg == "count" else f"{agg}({y_col})"

            if use_group:
                # Collect all unique x categories and groups
                all_cats_raw = result[x_col].cast(pl.Utf8).unique().sort().to_list()
                all_groups_raw = result[group_col].cast(pl.Utf8).unique().sort().to_list()
                # Build lookup
                lookup = {}
                for row in result.iter_rows(named=True):
                    lookup[(str(row[x_col]), str(row[group_col]))] = row["value"]
                # Sort categories by total value descending / ascending / label
                cat_totals = {c: sum(lookup.get((c, g), 0) or 0 for g in all_groups_raw) for c in all_cats_raw}
                if sort == "desc":
                    all_cats = sorted(all_cats_raw, key=lambda c: cat_totals[c], reverse=True)
                elif sort == "asc":
                    all_cats = sorted(all_cats_raw, key=lambda c: cat_totals[c])
                else:
                    all_cats = sorted(all_cats_raw)
                series = [
                    {"label": g, "values": [round(float(lookup.get((c, g), 0) or 0), 4) for c in all_cats]}
                    for g in all_groups_raw
                ]
                return {"grouped": True, "categories": all_cats, "series": series,
                        "x_col": x_col, "y_label": y_label}
            else:
                if sort == "desc":
                    result = result.sort("value", descending=True)
                elif sort == "asc":
                    result = result.sort("value", descending=False)
                else:
                    result = result.sort(x_col)
                cats = [str(v) for v in result[x_col].cast(pl.Utf8).to_list()]
                vals = [round(float(v), 4) if v is not None else 0 for v in result["value"].to_list()]
                return {"grouped": False, "categories": cats, "values": vals,
                        "x_col": x_col, "y_label": y_label}

        if body.block_type == "corr_matrix":
            import numpy as np
            numeric_types = (pl.Float64, pl.Float32, pl.Int64, pl.Int32, pl.Int16, pl.Int8,
                             pl.UInt64, pl.UInt32, pl.UInt16, pl.UInt8)
            all_numeric = [c for c in df.columns if df[c].dtype in numeric_types]
            requested = [c for c in (body.params.get("columns") or []) if c in all_numeric]
            numeric_cols = requested if requested else all_numeric
            if len(numeric_cols) < 2:
                raise HTTPException(status_code=400, detail="Need at least 2 numeric columns")
            sub = df.select(numeric_cols).drop_nulls()
            if sub.height < 2:
                raise HTTPException(status_code=400, detail="Not enough non-null rows to compute correlations")
            arr = sub.to_numpy().astype(float)
            corr = np.corrcoef(arr.T)
            n = len(numeric_cols)
            matrix = [[i, j, round(float(corr[i][j]), 4)] for i in range(n) for j in range(n)]
            result: dict = {"columns": numeric_cols, "matrix": matrix}
            if body.params.get("display") == "scatter":
                max_pts = int(body.params.get("max_points", 300))
                sample = sub.sample(n=min(max_pts, sub.height), shuffle=True, seed=42)
                result["raw_data"] = {col: sample[col].to_list() for col in numeric_cols}
            return result

        if body.block_type == "feature_importance":
            import numpy as np
            from scipy import stats as _stats
            numeric_types = (pl.Float64, pl.Float32, pl.Int64, pl.Int32, pl.Int16, pl.Int8,
                             pl.UInt64, pl.UInt32, pl.UInt16, pl.UInt8)

            y_col      = body.params.get("y_col", "")
            model_name = body.params.get("model", "decision_tree")
            if not y_col or y_col not in df.columns:
                raise HTTPException(status_code=400, detail="Select a target (Y) column")

            requested_x = [c for c in (body.params.get("columns") or []) if c in df.columns and c != y_col]
            x_cols = requested_x if requested_x else [c for c in df.columns if c != y_col]
            if not x_cols:
                raise HTTPException(status_code=400, detail="No feature columns available")

            # ── Detect regression vs classification ───────────────────────────
            y_raw = df[y_col].drop_nulls().cast(pl.Float64).to_list()
            unique_y = sorted(set(y_raw))
            is_classif = len(unique_y) <= 15 and all(v == int(v) for v in unique_y)

            if is_classif and len(unique_y) < 2:
                raise HTTPException(status_code=400, detail="Target column must have at least 2 distinct classes")

            # ── Build feature matrix (numeric + label-encoded categorical) ────
            sub_all = df.select(x_cols + [y_col]).drop_nulls()
            X_parts, feature_names = [], []
            for col in x_cols:
                if sub_all[col].dtype in numeric_types:
                    X_parts.append(sub_all[col].cast(pl.Float64).to_numpy().reshape(-1, 1))
                else:
                    X_parts.append(sub_all[col].cast(pl.Categorical).to_physical().cast(pl.Float64).to_numpy().reshape(-1, 1))
                feature_names.append(col)

            if not X_parts:
                raise HTTPException(status_code=400, detail="No usable feature columns")

            X = np.column_stack(X_parts)
            y_vec = sub_all[y_col].cast(pl.Float64).to_numpy()

            if is_classif:
                cls_list = sorted(set(int(v) for v in y_vec))
                cls_map  = {c: i for i, c in enumerate(cls_list)}
                inv_map  = {i: c for c, i in cls_map.items()}
                y_fit    = np.array([cls_map[int(v)] for v in y_vec])
            else:
                y_fit = y_vec

            # ── Fit model ─────────────────────────────────────────────────────
            if model_name == "linear":
                if is_classif:
                    # Per-feature eta² + overall linear fit
                    results = []
                    for i, col in enumerate(feature_names):
                        x_c = X[:, i]
                        if df[col].dtype in numeric_types:
                            r, p = _stats.pearsonr(x_c, y_vec)
                            results.append({"feature": col, "importance": round(r**2, 4), "r": round(float(r), 4), "p_value": round(float(p), 6), "type": "numeric"})
                        else:
                            y_mean = float(np.mean(y_vec))
                            ss_tot = float(np.sum((y_vec - y_mean)**2))
                            if ss_tot == 0: continue
                            groups = {}
                            for xi, yi in zip(x_c, y_vec):
                                groups.setdefault(xi, []).append(yi)
                            ss_bet = sum(len(g) * (np.mean(g) - y_mean)**2 for g in groups.values())
                            results.append({"feature": col, "importance": round(ss_bet / ss_tot, 4), "r": None, "p_value": None, "type": "categorical"})
                    results.sort(key=lambda x: x["importance"], reverse=True)
                    # linear model fit
                    X_aug = np.column_stack([X, np.ones(len(X))])
                    coef, _, _, _ = np.linalg.lstsq(X_aug, y_vec, rcond=None)
                    y_pred = X_aug @ coef
                    classes = np.array(sorted(set(int(v) for v in y_vec)))
                    y_pred_cls = classes[np.argmin(np.abs(classes[:, None] - y_pred[None, :]), axis=0)]
                    acc = float(np.mean(y_pred_cls == y_vec.astype(int)))
                    precs = []
                    for cls in classes:
                        tp = float(np.sum((y_pred_cls == cls) & (y_vec.astype(int) == cls)))
                        pp = float(np.sum(y_pred_cls == cls))
                        if pp > 0: precs.append(tp / pp)
                    model_fit = {"task": "classification", "accuracy": round(acc, 4), "precision": round(float(np.mean(precs)) if precs else 0.0, 4), "n_classes": len(classes)}
                else:
                    results = []
                    for i, col in enumerate(feature_names):
                        r, p = _stats.pearsonr(X[:, i], y_vec)
                        results.append({"feature": col, "importance": round(r**2, 4), "r": round(float(r), 4), "p_value": round(float(p), 6), "type": "numeric"})
                    results.sort(key=lambda x: x["importance"], reverse=True)
                    X_aug = np.column_stack([X, np.ones(len(X))])
                    coef, _, _, _ = np.linalg.lstsq(X_aug, y_vec, rcond=None)
                    y_pred = X_aug @ coef
                    ss_res = float(np.sum((y_vec - y_pred)**2))
                    ss_tot = float(np.sum((y_vec - float(np.mean(y_vec)))**2))
                    model_fit = {"task": "regression", "r2": round(1 - ss_res / ss_tot, 4) if ss_tot > 0 else 0.0, "rmse": round(float(np.sqrt(ss_res / len(y_vec))), 6)}
            else:
                try:
                    if model_name == "decision_tree":
                        from sklearn.tree import DecisionTreeRegressor, DecisionTreeClassifier
                        mdl = DecisionTreeClassifier(max_depth=8, random_state=42) if is_classif else DecisionTreeRegressor(max_depth=8, random_state=42)
                    elif model_name == "random_forest":
                        from sklearn.ensemble import RandomForestRegressor, RandomForestClassifier
                        mdl = RandomForestClassifier(n_estimators=100, random_state=42, n_jobs=-1) if is_classif else RandomForestRegressor(n_estimators=100, random_state=42, n_jobs=-1)
                    elif model_name == "xgboost":
                        from xgboost import XGBRegressor, XGBClassifier
                        mdl = XGBClassifier(n_estimators=100, random_state=42, n_jobs=-1, verbosity=0, eval_metric="logloss") if is_classif else XGBRegressor(n_estimators=100, random_state=42, n_jobs=-1, verbosity=0)
                    else:
                        raise HTTPException(status_code=400, detail=f"Unknown model: {model_name}")
                except ImportError as e:
                    raise HTTPException(status_code=400, detail=f"Package not installed: {e}. Run: pip install scikit-learn xgboost")

                mdl.fit(X, y_fit)
                y_pred_raw = mdl.predict(X)

                if is_classif:
                    y_pred_orig = np.array([inv_map.get(int(v), int(v)) for v in y_pred_raw])
                    y_orig      = y_vec.astype(int)
                    acc  = float(np.mean(y_pred_orig == y_orig))
                    precs = []
                    for cls in cls_list:
                        tp = float(np.sum((y_pred_orig == cls) & (y_orig == cls)))
                        pp = float(np.sum(y_pred_orig == cls))
                        if pp > 0: precs.append(tp / pp)
                    model_fit = {"task": "classification", "accuracy": round(acc, 4), "precision": round(float(np.mean(precs)) if precs else 0.0, 4), "n_classes": len(cls_list)}
                else:
                    ss_res = float(np.sum((y_vec - y_pred_raw)**2))
                    ss_tot = float(np.sum((y_vec - float(np.mean(y_vec)))**2))
                    model_fit = {"task": "regression", "r2": round(1 - ss_res / ss_tot, 4) if ss_tot > 0 else 0.0, "rmse": round(float(np.sqrt(ss_res / len(y_vec))), 6)}

                imps = mdl.feature_importances_
                results = [{"feature": feature_names[i], "importance": round(float(imps[i]), 4), "r": None, "p_value": None, "type": "model"} for i in range(len(feature_names))]
                results.sort(key=lambda x: x["importance"], reverse=True)

            return {"features": results, "y_col": y_col, "model_fit": model_fit, "model": model_name}

        if body.block_type == "predict_model":
            import numpy as np
            numeric_types = (pl.Float64, pl.Float32, pl.Int64, pl.Int32, pl.Int16, pl.Int8,
                             pl.UInt64, pl.UInt32, pl.UInt16, pl.UInt8)

            y_col      = body.params.get("y_col", "")
            model_name = body.params.get("model", "decision_tree")
            use_split  = str(body.params.get("use_split", "false")).lower() == "true"
            split_col  = (body.params.get("split_col") or "split").strip()

            if not y_col or y_col not in df.columns:
                raise HTTPException(status_code=400, detail="Select a target (Y) column")

            requested_x = [c for c in (body.params.get("columns") or []) if c in df.columns and c != y_col and c != split_col]
            x_cols = requested_x if requested_x else [c for c in df.columns if c != y_col and c != split_col]
            if not x_cols:
                raise HTTPException(status_code=400, detail="No feature columns available")

            # Build non-null subset; include split_col temporarily if needed
            extra_cols = [split_col] if (use_split and split_col in df.columns) else []
            sub_idx = df.select(x_cols + [y_col] + extra_cols).with_row_index("__idx__").filter(
                pl.all_horizontal([pl.col(c).is_not_null() for c in x_cols + [y_col]])
            )
            row_indices_all = sub_idx["__idx__"].to_list()
            sub = sub_idx.drop("__idx__")

            if sub.height == 0:
                raise HTTPException(status_code=400, detail="No rows remain after dropping nulls")

            # Determine train / test local index sets
            if use_split and split_col in df.columns and split_col in sub.columns:
                split_vals = sub[split_col].cast(pl.Utf8).to_list()
                train_local = [i for i, v in enumerate(split_vals) if v == "train"]
                test_local  = [i for i, v in enumerate(split_vals) if v == "test"]
                if not train_local:
                    raise HTTPException(status_code=400, detail="No 'train' rows found — run Data Split first")
                if not test_local:
                    raise HTTPException(status_code=400, detail="No 'test' rows found — run Data Split first")
                sub = sub.drop(split_col)
                is_split = True
            else:
                train_local = list(range(sub.height))
                test_local  = None
                is_split    = False

            if len(train_local) < 2:
                raise HTTPException(status_code=400, detail="Need at least 2 train rows")

            y_raw    = sub[y_col].cast(pl.Float64).to_list()
            unique_y = sorted(set(y_raw))
            is_classif = len(unique_y) <= 15 and all(v == int(v) for v in unique_y)

            if is_classif and len(unique_y) < 2:
                raise HTTPException(status_code=400, detail="Target column must have at least 2 distinct classes")

            X_parts = []
            for col in x_cols:
                if sub[col].dtype in numeric_types:
                    X_parts.append(sub[col].cast(pl.Float64).to_numpy().reshape(-1, 1))
                else:
                    X_parts.append(sub[col].cast(pl.Categorical).to_physical().cast(pl.Float64).to_numpy().reshape(-1, 1))
            X     = np.column_stack(X_parts)
            y_vec = sub[y_col].cast(pl.Float64).to_numpy()

            X_train = X[train_local]
            y_train = y_vec[train_local]
            X_test  = X[test_local]  if test_local else X
            y_test  = y_vec[test_local] if test_local else y_vec

            if is_classif:
                cls_list     = sorted(set(int(v) for v in y_vec))
                cls_map      = {c: i for i, c in enumerate(cls_list)}
                inv_map      = {i: c for c, i in cls_map.items()}
                y_fit_train  = np.array([cls_map[int(v)] for v in y_train])
            else:
                y_fit_train  = y_train

            try:
                if model_name == "linear":
                    if is_classif:
                        from sklearn.linear_model import LogisticRegression
                        mdl = LogisticRegression(max_iter=1000, random_state=42)
                    else:
                        from sklearn.linear_model import LinearRegression
                        mdl = LinearRegression()
                elif model_name == "decision_tree":
                    from sklearn.tree import DecisionTreeClassifier, DecisionTreeRegressor
                    mdl = DecisionTreeClassifier(max_depth=8, random_state=42) if is_classif else DecisionTreeRegressor(max_depth=8, random_state=42)
                elif model_name == "random_forest":
                    from sklearn.ensemble import RandomForestClassifier, RandomForestRegressor
                    mdl = RandomForestClassifier(n_estimators=100, random_state=42, n_jobs=-1) if is_classif else RandomForestRegressor(n_estimators=100, random_state=42, n_jobs=-1)
                elif model_name == "xgboost":
                    from xgboost import XGBClassifier, XGBRegressor
                    mdl = XGBClassifier(n_estimators=100, random_state=42, n_jobs=-1, verbosity=0, eval_metric="logloss") if is_classif else XGBRegressor(n_estimators=100, random_state=42, n_jobs=-1, verbosity=0)
                else:
                    raise HTTPException(status_code=400, detail=f"Unknown model: {model_name}")
            except ImportError as e:
                raise HTTPException(status_code=400, detail=f"Package not installed: {e}. Run: pip install scikit-learn xgboost")

            mdl.fit(X_train, y_fit_train)

            def _eval(X_ev, y_ev):
                yp_raw = mdl.predict(X_ev)
                if is_classif:
                    yp  = [int(inv_map.get(int(v), int(v))) for v in yp_raw]
                    ya  = [int(v) for v in y_ev]
                    from sklearn.metrics import accuracy_score, precision_score, recall_score, f1_score, confusion_matrix as sk_cm
                    cm = sk_cm(ya, yp, labels=cls_list).tolist()
                    return {
                        "task": "classification",
                        "accuracy":  round(float(accuracy_score(ya, yp)), 4),
                        "precision": round(float(precision_score(ya, yp, average="weighted", zero_division=0)), 4),
                        "recall":    round(float(recall_score(ya, yp, average="weighted", zero_division=0)), 4),
                        "f1":        round(float(f1_score(ya, yp, average="weighted", zero_division=0)), 4),
                        "classes": cls_list, "confusion_matrix": cm,
                    }, yp, ya
                else:
                    yp  = [round(float(v), 6) for v in yp_raw]
                    ya  = [float(v) for v in y_ev]
                    from sklearn.metrics import r2_score, mean_squared_error, mean_absolute_error
                    return {
                        "task": "regression",
                        "r2":   round(float(r2_score(ya, yp)), 4),
                        "rmse": round(float(np.sqrt(mean_squared_error(ya, yp))), 6),
                        "mae":  round(float(mean_absolute_error(ya, yp)), 6),
                    }, yp, ya

            train_fit, train_preds, train_actuals = _eval(X_train, y_train)
            if is_split:
                test_fit, test_preds, test_actuals = _eval(X_test, y_test)
                model_fit = test_fit
            else:
                test_fit, test_preds, test_actuals = None, None, None
                model_fit = train_fit

            train_row_indices = [row_indices_all[i] for i in train_local]
            test_row_indices  = [row_indices_all[i] for i in test_local] if test_local else None

            return {
                "y_col":             y_col,
                "model":             model_name,
                "is_classif":        is_classif,
                "is_split":          is_split,
                "model_fit":         model_fit,
                "model_fit_train":   train_fit   if is_split else None,
                "predictions":       test_preds  if is_split else train_preds,
                "actuals":           test_actuals if is_split else train_actuals,
                "row_indices":       test_row_indices if is_split else train_row_indices,
                "predictions_train": train_preds  if is_split else None,
                "actuals_train":     train_actuals if is_split else None,
                "row_indices_train": train_row_indices if is_split else None,
                "n_samples":         len(train_local),
                "n_test":            len(test_local) if test_local else None,
            }

        if body.block_type == "data_split":
            from sklearn.model_selection import train_test_split
            test_size    = max(0.05, min(0.95, float(body.params.get("test_size", 20)) / 100.0))
            seed         = int(body.params.get("random_seed", 42))
            stratify_col = body.params.get("stratify_col", "") or ""

            n = df.height
            if n < 2:
                raise HTTPException(status_code=400, detail="Need at least 2 rows to split")

            indices = list(range(n))
            strat = None
            if stratify_col and stratify_col in df.columns:
                strat = df[stratify_col].cast(pl.Utf8).to_list()

            try:
                train_idx, test_idx = train_test_split(
                    indices, test_size=test_size, random_state=seed, stratify=strat
                )
            except ImportError:
                raise HTTPException(status_code=400, detail="scikit-learn is required. Run: pip install scikit-learn")

            split_labels = ["train"] * n
            for i in test_idx:
                split_labels[i] = "test"

            return {
                "n_total":         n,
                "n_train":         len(train_idx),
                "n_test":          len(test_idx),
                "test_pct_actual": round(len(test_idx) / n * 100, 1),
                "split_labels":    split_labels,
            }

        if body.block_type == "value_count":
            column = body.params.get("column", "")
            top_n  = int(body.params.get("top_n", 20))
            dropna = str(body.params.get("dropna", "false")).lower() == "true"
            if not column:
                raise ValueError("column is required")
            if column not in df.columns:
                raise ValueError(f"Column '{column}' not found in data")
            total = len(df)
            s = df[column]
            if dropna:
                vc = s.drop_nulls().value_counts(sort=True)
                rows = [{"value": stringify_value(r[column]), "count": int(r["count"])}
                        for r in vc.to_dicts()]
            else:
                null_count = int(s.is_null().sum())
                vc = s.drop_nulls().value_counts(sort=True)
                rows = [{"value": stringify_value(r[column]), "count": int(r["count"])}
                        for r in vc.to_dicts()]
                if null_count > 0:
                    rows.append({"value": None, "count": null_count})
                rows.sort(key=lambda x: x["count"], reverse=True)
            return {
                "column": column,
                "total":  total,
                "values": rows[:top_n],
            }

        raise HTTPException(status_code=400, detail=f"Unknown block type: {body.block_type}")

    try:
        return await asyncio.to_thread(_run)
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


class PyExecRequest(BaseModel):
    code: str
    columns: list[str] = []
    rows: list[list] = []
    bq_schema: list[dict] = []


@app.post("/api/python/exec")
async def python_exec(body: PyExecRequest):
    def _run():
        import io, traceback, contextlib
        import pandas as pd
        import polars as pl

        if body.columns and body.rows:
            col_data = {col: [to_df_value(body.rows[r][i]) for r in range(len(body.rows))]
                        for i, col in enumerate(body.columns)}
            pl_df = _make_df(col_data)
            pl_df = _cast_df_by_bq_schema(pl_df, body.bq_schema)
            try:
                df = pl_df.to_pandas()
            except Exception:
                df = pd.DataFrame(col_data)
        else:
            df = pd.DataFrame()

        namespace: dict = {"df": df, "pd": pd, "pl": pl}

        captured = io.StringIO()
        error = None
        try:
            with contextlib.redirect_stdout(captured):
                exec(compile(body.code, "<python_block>", "exec"), namespace)
        except Exception:
            error = traceback.format_exc()

        return {"output": captured.getvalue(), "error": error}

    try:
        return await asyncio.to_thread(_run)
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


class PyFormatRequest(BaseModel):
    code: str


@app.post("/api/python/format")
async def python_format(body: PyFormatRequest):
    def _fmt():
        try:
            import autopep8
            return {"code": autopep8.fix_code(body.code, options={"max_line_length": 120}), "formatter": "autopep8"}
        except ImportError:
            pass
        try:
            import black
            return {"code": black.format_str(body.code, mode=black.Mode()), "formatter": "black"}
        except ImportError:
            pass
        return {"code": body.code, "error": "Install autopep8 or black: pip install autopep8"}
    try:
        return await asyncio.to_thread(_fmt)
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


import platform as _platform
import shutil as _shutil
import subprocess as _subprocess
import re as _re

# ── Filesystem API ─────────────────────────────────────────────────────────────

def _safe_table_name(name: str) -> str:
    safe = _re.sub(r'[^a-zA-Z0-9_]', '_', name.strip()).lower()
    if not safe or safe[0].isdigit():
        safe = 's_' + safe
    return safe

def _read_excel_sheets(path):
    """Read all sheets from an xlsx file using openpyxl.
    Returns dict: {sheet_name: DataFrame}"""
    import polars as pl
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheets = {}
    for sname in wb.sheetnames:
        ws = wb[sname]
        all_rows = list(ws.values)
        if not all_rows:
            sheets[sname] = pl.DataFrame()
            continue
        raw_headers = all_rows[0]
        headers, seen = [], {}
        for i, h in enumerate(raw_headers):
            base = str(h).strip() if h is not None else f'col_{i}'
            if base in seen:
                seen[base] += 1
                headers.append(f'{base}_{seen[base]}')
            else:
                seen[base] = 0
                headers.append(base)
        data = all_rows[1:]
        col_data = {
            h: [str(r[i]) if i < len(r) and r[i] is not None else None for r in data]
            for i, h in enumerate(headers)
        }
        df = pl.DataFrame(col_data, infer_schema_length=min(1000, len(data)) or None)
        sheets[sname] = df
    wb.close()
    return sheets

class FsListReq(BaseModel):
    path: str

class FsWriteReq(BaseModel):
    path: str
    content: str

class FsRenameReq(BaseModel):
    old_path: str
    new_path: str

class FsMoveReq(BaseModel):
    src: str
    dst: str

class FsMkdirReq(BaseModel):
    path: str

class FsDeleteReq(BaseModel):
    path: str


def _fsp(p: str) -> Path:
    return Path(os.path.expanduser(p)).resolve()


@app.post("/api/fs/list")
async def fs_list(req: FsListReq):
    p = _fsp(req.path)
    if not p.exists() or not p.is_dir():
        raise HTTPException(status_code=400, detail=f"Not a directory: {p}")
    items = []
    try:
        for child in sorted(p.iterdir(), key=lambda x: (not x.is_dir(), x.name.lower())):
            if child.name.startswith('.'):
                continue
            items.append({
                "name": child.name,
                "path": str(child),
                "is_dir": child.is_dir(),
                "ext": child.suffix.lower() if child.is_file() else "",
            })
    except PermissionError:
        pass
    return {"items": items, "path": str(p)}


@app.get("/api/fs/read")
async def fs_read(path: str):
    p = _fsp(path)
    if not p.exists() or not p.is_file():
        raise HTTPException(status_code=400, detail="File not found")
    try:
        content = p.read_text(encoding="utf-8")
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))
    return {"content": content, "path": str(p)}


@app.post("/api/fs/write")
async def fs_write(req: FsWriteReq):
    p = _fsp(req.path)
    p.parent.mkdir(parents=True, exist_ok=True)
    p.write_text(req.content, encoding="utf-8")
    return {"ok": True, "path": str(p)}


@app.post("/api/fs/rename")
async def fs_rename(req: FsRenameReq):
    old = _fsp(req.old_path)
    new = _fsp(req.new_path)
    if not old.exists():
        raise HTTPException(status_code=400, detail="Source not found")
    if new.exists():
        raise HTTPException(status_code=400, detail="Destination already exists")
    old.rename(new)
    return {"ok": True, "new_path": str(new)}


@app.post("/api/fs/move")
async def fs_move(req: FsMoveReq):
    src = _fsp(req.src)
    dst = _fsp(req.dst)
    if not src.exists():
        raise HTTPException(status_code=400, detail="Source not found")
    if dst.is_dir():
        dst = dst / src.name
    if dst.exists():
        raise HTTPException(status_code=400, detail="Destination already exists")
    _shutil.move(str(src), str(dst))
    return {"ok": True, "new_path": str(dst)}


@app.post("/api/fs/mkdir")
async def fs_mkdir(req: FsMkdirReq):
    p = _fsp(req.path)
    p.mkdir(parents=True, exist_ok=True)
    return {"ok": True, "path": str(p)}


@app.get("/api/fs/read-data")
async def fs_read_data(path: str):
    p = _fsp(path)
    if not p.exists():
        raise HTTPException(status_code=400, detail="File not found")
    ext = p.suffix.lower()
    if ext not in ('.csv', '.tsv', '.json', '.jsonl', '.parquet', '.xlsx', '.xls'):
        raise HTTPException(status_code=400, detail=f"Unsupported data file type: {ext}")
    def _read():
        import polars as pl
        if ext in ('.xlsx', '.xls'):
            sheets_frames = _read_excel_sheets(p)
            if not sheets_frames:
                return {"columns": [], "schema": [], "rows": [], "totalRows": 0, "sheets": {}}
            sheets_out = {}
            first_df, first_name = None, None
            for sname, sdf in sheets_frames.items():
                total = sdf.height
                trunc_df = sdf.head(MAX_ROWS) if total > MAX_ROWS else sdf
                ss = [{"name": c, "type": str(trunc_df.schema[c])} for c in trunc_df.columns]
                tname = _safe_table_name(sname)
                sheets_out[sname] = {
                    "tableName": tname,
                    "columns": trunc_df.columns,
                    "schema": ss,
                    "rows": [[stringify_value(v) for v in row] for row in trunc_df.rows()],
                    "totalRows": total,
                }
                if first_df is None:
                    first_df, first_name = trunc_df, sname
            s = sheets_out[first_name]
            return {
                "columns": s["columns"], "schema": s["schema"],
                "rows": s["rows"], "totalRows": sheets_out[first_name]["totalRows"],
                "sheets": sheets_out,
            }
        # Single-sheet formats
        if ext in ('.csv', '.tsv'):
            df = pl.read_csv(p, separator='\t' if ext == '.tsv' else ',', infer_schema_length=1000)
        elif ext == '.json':
            df = pl.read_json(p)
        elif ext == '.jsonl':
            df = pl.read_ndjson(p)
        else:
            df = pl.read_parquet(p)
        total = df.height
        if total > MAX_ROWS:
            df = df.head(MAX_ROWS)
        schema = [{"name": c, "type": str(df.schema[c])} for c in df.columns]
        return {
            "columns": df.columns,
            "schema": schema,
            "rows": [[stringify_value(v) for v in row] for row in df.rows()],
            "totalRows": total,
            "sheets": None,
        }
    try:
        return await asyncio.to_thread(_read)
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@app.post("/api/fs/delete")
async def fs_delete(req: FsDeleteReq):
    p = _fsp(req.path)
    if not p.exists():
        raise HTTPException(status_code=400, detail="Path not found")
    if p.is_dir():
        _shutil.rmtree(p)
    else:
        p.unlink()
    return {"ok": True}


@app.get("/api/fs/pick-folder")
async def fs_pick_folder():
    def _pick():
        if _platform.system() == "Darwin":
            r = _subprocess.run(
                ["osascript", "-e", "POSIX path of (choose folder)"],
                capture_output=True, text=True,
            )
            return r.stdout.strip() if r.returncode == 0 else None
        else:
            try:
                import tkinter as tk
                from tkinter import filedialog
                root = tk.Tk()
                root.withdraw()
                root.attributes("-topmost", True)
                path = filedialog.askdirectory(title="Select Folder")
                root.destroy()
                return path or None
            except Exception:
                return None

    path = await asyncio.to_thread(_pick)
    return {"path": path}


app.mount("/", StaticFiles(directory=str(STATIC_DIR), html=True), name="static")
